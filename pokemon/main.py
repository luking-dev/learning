from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.alignment import Alignment

import requests
import io

wb = Workbook()
ws = wb.active

poke_url = "https://pokeapi.co/api/v2/pokemon/?limit=151"

res = requests.get(poke_url).json()["results"]

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15

try:
    for i in range(len(res)):
        name = res[i]["name"].capitalize()

        img_url = requests.get(res[i]["url"]).json()["sprites"]["front_default"]

        img_res = requests.get(img_url)
        
        img_file = io.BytesIO(img_res.content)
        img = Image(img_file)

        ws.row_dimensions[i + 1].height = 70

        ws["A" + str(i + 1)] = f"#{i + 1}"
        ws["A" + str(i + 1)].alignment = Alignment(horizontal="center", vertical="center")

        ws["B" + str(i + 1)] = name
        ws["B" + str(i + 1)].alignment = Alignment(horizontal="center", vertical="center")

        ws.add_image(img, "C" + str(i + 1))
    
    wb.save("pokemons.xlsx")
    raise Exception("Solicitud exitosa. Proceso completado.")
except Exception as e:
    print(e)